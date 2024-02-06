import openpyxl
import pandas as pd


def openpyxl_read_write():
    # https://www.datacamp.com/tutorial/python-excel-tutorial
    # https://www.geeksforgeeks.org/excel-automation-with-openpyxl-in-python/
    filename = "data.xlsx"
    wb = openpyxl.load_workbook(filename)  # load excel workbook
    print(wb.sheetnames)  # print sheetnames
    # print(dir(wb))  # Workbook object: https://openpyxl.readthedocs.io/en/latest/api/openpyxl.workbook.workbook.html

    list1 = wb["List1"]  # pick sheet named List1
    # print(dir(list1))  # Worksheet object: https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html
    print(list(list1.values))  # print its content

    print(list1.cell(row=1, column=2).value)  # print content of cell "B1" (indexing starts from 1)
    list1.cell(row=1, column=8).value = "modified"  # modify cell "H1"

    wb.save(filename=filename)  # save the file


def pandas_read_write():
    df = pd.read_excel("data.xlsx", index_col=0)
    print(df)

    df["SUM"] = df.sum(axis=1)
    # print(df.corr())
    # df.corr().to_excel("data2.xlsx", index=False)  # write to a new file

    # Excel Writer object - append mode
    with pd.ExcelWriter("data.xlsx", mode="a") as writer:
        df.to_excel(writer, sheet_name="List2")
        df.corr().to_excel(writer, sheet_name="Corr")


if __name__ == "__main__":
    openpyxl_read_write()
    # pandas_read_write()
